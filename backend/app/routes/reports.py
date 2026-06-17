"""Rutas para reportes y exportación de datos"""
from flask import Blueprint, jsonify, send_file, request
from datetime import datetime
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)

# Datos simulados para pruebas
mock_camas = {
    1: {
        'urgencias': [
            {'id': 1, 'numero': 'U-001', 'disponible': True},
            {'id': 2, 'numero': 'U-002', 'disponible': False},
            {'id': 3, 'numero': 'U-003', 'disponible': True},
            {'id': 4, 'numero': 'U-004', 'disponible': True},
            {'id': 5, 'numero': 'U-005', 'disponible': False},
        ],
        'pediatria': [
            {'id': 6, 'numero': 'P-001', 'disponible': True},
            {'id': 7, 'numero': 'P-002', 'disponible': True},
            {'id': 8, 'numero': 'P-003', 'disponible': False},
            {'id': 9, 'numero': 'P-004', 'disponible': True},
            {'id': 10, 'numero': 'P-005', 'disponible': True},
            {'id': 11, 'numero': 'P-006', 'disponible': True},
            {'id': 12, 'numero': 'P-007', 'disponible': False},
            {'id': 13, 'numero': 'P-008', 'disponible': True},
        ],
        'cirugia': [
            {'id': 14, 'numero': 'C-001', 'disponible': True},
            {'id': 15, 'numero': 'C-002', 'disponible': True},
            {'id': 16, 'numero': 'C-003', 'disponible': False},
            {'id': 17, 'numero': 'C-004', 'disponible': True},
            {'id': 18, 'numero': 'C-005', 'disponible': True},
            {'id': 19, 'numero': 'C-006', 'disponible': True},
            {'id': 20, 'numero': 'C-007', 'disponible': False},
            {'id': 21, 'numero': 'C-008', 'disponible': True},
            {'id': 22, 'numero': 'C-009', 'disponible': True},
            {'id': 23, 'numero': 'C-010', 'disponible': True},
            {'id': 24, 'numero': 'C-011', 'disponible': True},
            {'id': 25, 'numero': 'C-012', 'disponible': False},
        ],
        'medicina_interna': [
            {'id': 26, 'numero': 'M-001', 'disponible': True},
            {'id': 27, 'numero': 'M-002', 'disponible': True},
            {'id': 28, 'numero': 'M-003', 'disponible': False},
            {'id': 29, 'numero': 'M-004', 'disponible': True},
            {'id': 30, 'numero': 'M-005', 'disponible': True},
            {'id': 31, 'numero': 'M-006', 'disponible': True},
            {'id': 32, 'numero': 'M-007', 'disponible': True},
            {'id': 33, 'numero': 'M-008', 'disponible': False},
            {'id': 34, 'numero': 'M-009', 'disponible': True},
            {'id': 35, 'numero': 'M-010', 'disponible': True},
            {'id': 36, 'numero': 'M-011', 'disponible': True},
            {'id': 37, 'numero': 'M-012', 'disponible': True},
            {'id': 38, 'numero': 'M-013', 'disponible': True},
            {'id': 39, 'numero': 'M-014', 'disponible': True},
            {'id': 40, 'numero': 'M-015', 'disponible': False},
            {'id': 41, 'numero': 'M-016', 'disponible': True},
            {'id': 42, 'numero': 'M-017', 'disponible': True},
            {'id': 43, 'numero': 'M-018', 'disponible': True},
            {'id': 44, 'numero': 'M-019', 'disponible': True},
            {'id': 45, 'numero': 'M-020', 'disponible': True},
        ]
    }
}

mock_refrigeradores = {
    1: [
        {'id': 1, 'nombre': 'Refrigerador Vacunas A', 'temperatura': 4.5, 'estado': 'OK'},
        {'id': 2, 'nombre': 'Refrigerador Vacunas B', 'temperatura': 7.8, 'estado': 'ALERTA'},
        {'id': 3, 'nombre': 'Congelador Banco Sangre', 'temperatura': -25.2, 'estado': 'OK'},
    ]
}

def calcular_estadisticas_camas(hospital_id):
    """Calcula estadísticas de camas para un hospital"""
    camas = mock_camas.get(hospital_id, {})
    total = 0
    disponibles = 0
    
    for piso, lista_camas in camas.items():
        for cama in lista_camas:
            total += 1
            if cama['disponible']:
                disponibles += 1
    
    return {
        'total': total,
        'disponibles': disponibles,
        'ocupadas': total - disponibles,
        'porcentaje_ocupacion': round((total - disponibles) / total * 100, 2) if total > 0 else 0
    }

@reports_bp.route('/api/reportes/camas/<int:hospital_id>', methods=['GET'])
def obtener_camas(hospital_id):
    """Obtiene estadísticas de camas por hospital"""
    try:
        stats = calcular_estadisticas_camas(hospital_id)
        camas_por_piso = {}
        
        for piso, lista_camas in mock_camas.get(hospital_id, {}).items():
            disponibles = sum(1 for c in lista_camas if c['disponible'])
            camas_por_piso[piso] = {
                'total': len(lista_camas),
                'disponibles': disponibles,
                'ocupadas': len(lista_camas) - disponibles
            }
        
        return jsonify({
            'total': stats['total'],
            'disponibles': stats['disponibles'],
            'ocupadas': stats['ocupadas'],
            'porcentaje_ocupacion': stats['porcentaje_ocupacion'],
            'por_piso': camas_por_piso,
            'ultima_actualizacion': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/api/reportes/temperatura/<int:hospital_id>', methods=['GET'])
def obtener_temperatura(hospital_id):
    """Obtiene temperatura de refrigeradores"""
    try:
        refrig = mock_refrigeradores.get(hospital_id, [])
        
        return jsonify({
            'refrigeradores': refrig,
            'temperatura_promedio': round(sum(r['temperatura'] for r in refrig) / len(refrig), 2) if refrig else 0,
            'ultima_actualizacion': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/api/reportes/descargar', methods=['POST'])
def generar_reporte_excel():
    """Genera un reporte en Excel"""
    try:
        data = request.json
        hospital_id = data.get('hospital_id', 1)
        tipo_reporte = data.get('tipo_reporte', 'ocupacion')
        fecha_inicio = data.get('fecha_inicio', '2026-06-01')
        fecha_fin = data.get('fecha_fin', '2026-06-16')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "REPORTE DE HOSPITAL"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Hospital ID: {hospital_id}"
        ws['A3'] = f"Período: {fecha_inicio} a {fecha_fin}"
        ws['A4'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Contenido según tipo
        if tipo_reporte == 'ocupacion':
            ws['A6'] = "Ocupación de Camas"
            ws['A6'].font = Font(bold=True, size=12)
            
            # Headers
            headers = ['Piso', 'Total', 'Disponibles', 'Ocupadas', 'Porcentaje Ocupación']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos de camas
            row = 8
            for piso, camas in mock_camas.get(hospital_id, {}).items():
                total = len(camas)
                disponibles = sum(1 for c in camas if c['disponible'])
                ocupadas = total - disponibles
                porcentaje = round(ocupadas / total * 100, 2) if total > 0 else 0
                
                datos = [piso, total, disponibles, ocupadas, f"{porcentaje}%"]
                for col, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = valor
                    cell.border = border
                    if col == 5:
                        cell.alignment = Alignment(horizontal='right')
                
                row += 1
            
            # Total
            row += 1
            ws.cell(row=row, column=1).value = "TOTAL"
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            stats = calcular_estadisticas_camas(hospital_id)
            ws.cell(row=row, column=2).value = stats['total']
            ws.cell(row=row, column=3).value = stats['disponibles']
            ws.cell(row=row, column=4).value = stats['ocupadas']
            ws.cell(row=row, column=5).value = f"{stats['porcentaje_ocupacion']}%"
            
        elif tipo_reporte == 'temperatura':
            ws['A6'] = "Temperatura de Refrigeradores"
            ws['A6'].font = Font(bold=True, size=12)
            
            # Headers
            headers = ['Nombre', 'Temperatura (°C)', 'Estado', 'Mín', 'Máx']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            row = 8
            for refrig in mock_refrigeradores.get(hospital_id, []):
                datos = [
                    refrig['nombre'],
                    refrig['temperatura'],
                    refrig['estado'],
                    2.0,
                    8.0
                ]
                for col, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = valor
                    cell.border = border
                    if col > 1:
                        cell.alignment = Alignment(horizontal='right')
                
                row += 1
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_{hospital_id}_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@reports_bp.route('/api/reportes/descargar-json', methods=['POST'])
def descargar_json():
    """Descarga datos en JSON"""
    try:
        import json as json_module
        data = request.json
        hospital_id = data.get('hospital_id', 1)
        
        datos = {
            'hospital_id': hospital_id,
            'fecha_generacion': datetime.now().isoformat(),
            'camas': calcular_estadisticas_camas(hospital_id),
            'refrigeradores': mock_refrigeradores.get(hospital_id, [])
        }
        
        output = io.BytesIO()
        output.write(json_module.dumps(datos).encode('utf-8'))
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_{hospital_id}_{timestamp}.json"
        
        return send_file(
            output,
            mimetype='application/json',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
